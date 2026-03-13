#!/usr/bin/env python3
"""Bridge the checkpoint logger CSV with the Celeste practice spreadsheet."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell import Cell


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_LOG_DIR = ROOT / "timer_data" / "checkpoint_logs"
DEFAULT_WORKBOOK = ROOT / "Celeste Any% Practice Sheet.xlsx"
DEFAULT_MAP = Path(__file__).resolve().with_name("checkpoint_sheet_map.json")
TARGET_SHEET = "Celeste Any% 6a  CPs"


@dataclass
class CsvEntry:
    key: str
    segment_ms: int
    end_time_iso: Optional[str]


@dataclass
class SheetRow:
    sheet_chapter: str
    sheet_cp: str
    time_cell: Cell
    date_cell: Cell
    time_is_formula: bool


@dataclass
class UpdatePlan:
    row: SheetRow
    key: str
    new_time: str
    new_date: Optional[dt.datetime]
    reason: str


class BridgeError(Exception):
    pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Update practice sheet PBs from the latest checkpoint log"
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write changes to the workbook (default: dry-run)",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="Path to Celeste Any%% Practice Sheet.xlsx",
    )
    parser.add_argument(
        "--log",
        type=Path,
        default=None,
        help="Path to a checkpoint CSV (default: latest in timer_data/checkpoint_logs)",
    )
    parser.add_argument(
        "--map",
        type=Path,
        default=DEFAULT_MAP,
        help="Path to checkpoint_sheet_map.json",
    )
    return parser.parse_args()


def load_mapping(map_path: Path) -> Dict[str, Dict[str, str]]:
    if not map_path.is_file():
        raise BridgeError(f"Mapping file not found: {map_path}")
    with map_path.open("r", encoding="utf-8") as fp:
        data = json.load(fp)
    if not isinstance(data, dict):
        raise BridgeError("Mapping file must contain a JSON object")
    return data


def find_latest_csv(log_dir: Path) -> Path:
    if not log_dir.is_dir():
        raise BridgeError(f"Log directory not found: {log_dir}")
    csv_files = sorted(log_dir.glob("*.csv"), reverse=True)
    if not csv_files:
        raise BridgeError(f"No CSV logs found in {log_dir}")
    return csv_files[0]


def normalize_checkpoint_name(name: str) -> str:
    return " ".join(name.strip().split())


def mode_to_letter(mode_value: str) -> str:
    try:
        mode_int = int(mode_value)
    except (TypeError, ValueError):
        return str(mode_value)
    return {0: "A", 1: "B", 2: "C"}.get(mode_int, str(mode_int))


def csv_key(row: Dict[str, str]) -> str:
    chapter = row.get("chapter")
    mode = row.get("mode")
    checkpoint_name = row.get("checkpoint_name")
    if chapter is None or mode is None or checkpoint_name is None:
        raise BridgeError("CSV row missing required columns")
    norm_name = normalize_checkpoint_name(checkpoint_name)
    letter = mode_to_letter(mode)
    return f"{chapter}:{letter}:{norm_name}"


def read_best_segments(csv_path: Path) -> Tuple[Dict[str, CsvEntry], Dict[str, str]]:
    best: Dict[str, CsvEntry] = {}
    invalid: Dict[str, str] = {}
    with csv_path.open("r", newline="") as fp:
        reader = csv.DictReader(fp)
        required = {
            "chapter",
            "mode",
            "checkpoint_name",
            "segment_ms",
            "end_time_iso",
        }
        if not required.issubset(reader.fieldnames or []):
            missing = required - set(reader.fieldnames or [])
            raise BridgeError(f"CSV missing columns: {', '.join(sorted(missing))}")
        for row in reader:
            key = csv_key(row)
            seg_raw = row.get("segment_ms")
            try:
                segment_ms = int(seg_raw)
            except (TypeError, ValueError):
                if key not in best:
                    invalid[key] = "segment_non_positive"
                continue
            if segment_ms <= 0:
                if key not in best:
                    invalid[key] = "segment_non_positive"
                continue
            entry = CsvEntry(key=key, segment_ms=segment_ms, end_time_iso=row.get("end_time_iso"))
            current = best.get(key)
            if current is None:
                best[key] = entry
                invalid.pop(key, None)
                continue
            if segment_ms < current.segment_ms:
                best[key] = entry
                continue
            if segment_ms == current.segment_ms:
                current_dt = parse_iso8601(current.end_time_iso)
                new_dt = parse_iso8601(entry.end_time_iso)
                if new_dt is not None and (current_dt is None or new_dt > current_dt):
                    best[key] = entry
    return best, invalid


def parse_iso8601(value: Optional[str]) -> Optional[dt.datetime]:
    if not value:
        return None
    try:
        return dt.datetime.fromisoformat(value)
    except ValueError:
        return None


def parse_time_cell(value: Optional[str]) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        if ":" in text:
            minutes_str, rest = text.split(":", 1)
            minutes = int(minutes_str)
            seconds = float(rest)
            total_seconds = minutes * 60 + seconds
        else:
            total_seconds = float(text)
    except ValueError:
        return None
    return int(round(total_seconds * 1000))


def format_segment_ms(segment_ms: int) -> str:
    minutes, remainder = divmod(segment_ms, 60000)
    seconds = remainder // 1000
    millis = remainder % 1000
    if minutes:
        return f"{minutes}:{seconds:02d}.{millis:03d}"
    return f"{seconds}.{millis:03d}"


def build_sheet_index(ws) -> Tuple[Dict[Tuple[str, str], SheetRow], set]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header_map = {str(cell.value).strip(): idx for idx, cell in enumerate(header_row) if cell.value}
    required_headers = {"Chapter", "CP", "Time", "Date"}
    if not required_headers.issubset(header_map):
        missing = required_headers - set(header_map)
        raise BridgeError(f"Sheet missing columns: {', '.join(sorted(missing))}")
    chap_idx = header_map["Chapter"]
    cp_idx = header_map["CP"]
    time_idx = header_map["Time"]
    date_idx = header_map["Date"]
    index: Dict[Tuple[str, str], SheetRow] = {}
    ambiguous: set = set()
    current_chapter = None
    for row in ws.iter_rows(min_row=2):
        chapter_cell = row[chap_idx]
        cp_cell = row[cp_idx]
        if chapter_cell.value not in (None, ""):
            current_chapter = str(chapter_cell.value).strip()
        if cp_cell.value in (None, "") or current_chapter in (None, ""):
            continue
        sheet_chapter = current_chapter
        sheet_cp = str(cp_cell.value).strip()
        key = (sheet_chapter, sheet_cp)
        time_cell = row[time_idx]
        date_cell = row[date_idx]
        entry = SheetRow(
            sheet_chapter=sheet_chapter,
            sheet_cp=sheet_cp,
            time_cell=time_cell,
            date_cell=date_cell,
            time_is_formula=time_cell.data_type == "f",
        )
        if key in index:
            ambiguous.add(key)
        index.setdefault(key, entry)
    return index, ambiguous


def format_date_value(value) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value) if value not in (None, "") else ""


def main() -> None:
    args = parse_args()
    try:
        mapping = load_mapping(args.map)
        if args.apply and args.log is None:
            raise BridgeError("--apply requires --log to specify the source CSV")
        csv_path = args.log or find_latest_csv(DEFAULT_LOG_DIR)
        best_segments, invalid_segments = read_best_segments(csv_path)
        if not best_segments and not invalid_segments:
            print("No valid segments found in CSV; nothing to do.")
            return
        if not args.workbook.is_file():
            raise BridgeError(f"Workbook not found: {args.workbook}")
        wb = load_workbook(args.workbook, data_only=False)
        if TARGET_SHEET not in wb.sheetnames:
            raise BridgeError(f"Sheet '{TARGET_SHEET}' not found in workbook")
        ws = wb[TARGET_SHEET]
        sheet_index, ambiguous_rows = build_sheet_index(ws)
        results: List[Dict[str, str]] = []
        updates: List[UpdatePlan] = []
        all_keys = sorted(set(best_segments.keys()) | set(invalid_segments.keys()))
        for key in all_keys:
            entry = best_segments.get(key)
            mapping_entry = mapping.get(key)
            if entry is None:
                reason = invalid_segments.get(key, "segment_non_positive")
                results.append(
                    {
                        "status": "SKIP",
                        "key": key,
                        "target": "-",
                        "old_time": "",
                        "new_time": "",
                        "old_date": "",
                        "new_date": "",
                        "reason": reason,
                    }
                )
                continue
            if mapping_entry is None:
                results.append(
                    {
                        "status": "SKIP",
                        "key": key,
                        "target": "-",
                        "old_time": "",
                        "new_time": format_segment_ms(entry.segment_ms),
                        "old_date": "",
                        "new_date": "",
                        "reason": "mapping_missing",
                    }
                )
                continue
            sheet_key = (mapping_entry["sheet_chapter"], mapping_entry["sheet_cp"])
            if sheet_key in ambiguous_rows:
                results.append(
                    {
                        "status": "SKIP",
                        "key": key,
                        "target": f"{sheet_key[0]} / {sheet_key[1]}",
                        "old_time": "",
                        "new_time": "",
                        "old_date": "",
                        "new_date": "",
                        "reason": "sheet_row_ambiguous",
                    }
                )
                continue
            row = sheet_index.get(sheet_key)
            if row is None:
                results.append(
                    {
                        "status": "SKIP",
                        "key": key,
                        "target": f"{sheet_key[0]} / {sheet_key[1]}",
                        "old_time": "",
                        "new_time": "",
                        "old_date": "",
                        "new_date": "",
                        "reason": "sheet_row_missing",
                    }
                )
                continue
            if row.time_is_formula:
                results.append(
                    {
                        "status": "SKIP",
                        "key": key,
                        "target": f"{sheet_key[0]} / {sheet_key[1]}",
                        "old_time": str(row.time_cell.value or ""),
                        "new_time": format_segment_ms(entry.segment_ms),
                        "old_date": format_date_value(row.date_cell.value),
                        "new_date": "",
                        "reason": "formula_cell",
                    }
                )
                continue
            current_ms = parse_time_cell(row.time_cell.value)
            if current_ms is not None and entry.segment_ms >= current_ms:
                results.append(
                    {
                        "status": "SKIP",
                        "key": key,
                        "target": f"{sheet_key[0]} / {sheet_key[1]}",
                        "old_time": str(row.time_cell.value or ""),
                        "new_time": format_segment_ms(entry.segment_ms),
                        "old_date": format_date_value(row.date_cell.value),
                        "new_date": format_date_value(row.date_cell.value),
                        "reason": "not_faster",
                    }
                )
                continue
            if current_ms is None and str(row.time_cell.value or "").strip() and parse_time_cell(row.time_cell.value) is None:
                results.append(
                    {
                        "status": "SKIP",
                        "key": key,
                        "target": f"{sheet_key[0]} / {sheet_key[1]}",
                        "old_time": str(row.time_cell.value or ""),
                        "new_time": format_segment_ms(entry.segment_ms),
                        "old_date": format_date_value(row.date_cell.value),
                        "new_date": format_date_value(row.date_cell.value),
                        "reason": "time_parse_failed",
                    }
                )
                continue
            new_time_text = format_segment_ms(entry.segment_ms)
            date_value = parse_iso8601(entry.end_time_iso)
            new_date_value: Optional[dt.datetime]
            reason = "improved"
            if date_value is None:
                new_date_value = None
                reason = "improved_no_date"
            else:
                new_date_value = dt.datetime.combine(date_value.date(), dt.time())
            updates.append(
                UpdatePlan(
                    row=row,
                    key=key,
                    new_time=new_time_text,
                    new_date=new_date_value,
                    reason=reason,
                )
            )
            results.append(
                {
                    "status": "UPDATE",
                    "key": key,
                    "target": f"{sheet_key[0]} / {sheet_key[1]}",
                    "old_time": str(row.time_cell.value or ""),
                    "new_time": new_time_text,
                    "old_date": format_date_value(row.date_cell.value),
                    "new_date": new_date_value.date().isoformat() if new_date_value else "(unchanged)",
                    "reason": "improved" if new_date_value else "improved_no_date",
                }
            )

        print_report(csv_path, args.workbook, args.apply, updates, results)

        if args.apply and updates:
            backup_path = args.workbook.with_name(
                f"{args.workbook.name}.bak-{dt.datetime.now().strftime('%Y%m%d-%H%M%S')}"
            )
            shutil.copy2(args.workbook, backup_path)
            for plan in updates:
                plan.row.time_cell.value = plan.new_time
                if plan.new_date is not None:
                    plan.row.date_cell.value = plan.new_date
            wb.save(args.workbook)
            print(f"Backup written to {backup_path}")
            print(f"Workbook updated: {args.workbook}")
        elif not updates:
            print("No improvements found.")
    except BridgeError as err:
        print(f"Error: {err}", file=sys.stderr)
        sys.exit(1)


def print_report(csv_path: Path, workbook_path: Path, apply: bool, updates, results) -> None:
    mode = "APPLY" if apply else "DRY-RUN"
    print(f"Mode: {mode}")
    print(f"CSV: {csv_path}")
    print(f"Workbook: {workbook_path}")
    headers = ["Status", "Logger Key", "Sheet Target", "Old Time", "New Time", "Old Date", "New Date", "Reason"]
    rows = [
        [
            r["status"],
            r["key"],
            r["target"],
            r["old_time"],
            r["new_time"],
            r["old_date"],
            r["new_date"],
            r["reason"],
        ]
        for r in results
    ]
    widths = [len(h) for h in headers]
    for row in rows:
        for idx, value in enumerate(row):
            widths[idx] = max(widths[idx], len(str(value)))

    def fmt_row(values):
        return " | ".join(str(val).ljust(widths[idx]) for idx, val in enumerate(values))

    print("-" * (sum(widths) + 3 * (len(headers) - 1)))
    print(fmt_row(headers))
    print("-" * (sum(widths) + 3 * (len(headers) - 1)))
    for row in rows:
        print(fmt_row(row))
    print("-" * (sum(widths) + 3 * (len(headers) - 1)))


if __name__ == "__main__":
    main()
